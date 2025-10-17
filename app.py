#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
App de análise fundamentalista usando getDataCVM (DFP).

- Padrões específicos para bancos
- Alinhamento por ano entre DRE e BPP
- "NA" em vez de 0 quando a conta não é encontrada
- Fallback para *_ind quando *_con vier vazio
- Busca mais robusta (e opção --exact)
- Exporta XLSX com abas DRE_ref e BPP_ref
- Flags de diagnóstico (--debug-accounts, --samples)

NOVO:
- Extração de séries históricas por ano (DRE & BPP) para várias contas-chave
- Indicadores completos: ROA, ROE, Margem Bruta, Margem Operacional (EBIT), Margem Líquida,
  Endividamento, Composição da Dívida, Liquidez Corrente e Seca
- Crescimento (YoY) de Receita e Lucro; cálculo em todas as janelas disponíveis
- Nova aba no Excel: 'indicadores_hist' (toda a série de indicadores por ano)
- Nova aba no Excel: 'contas_hist' (as contas extraídas por ano, para auditoria)

excel:
- Aba 'resumo' simplificada e legível (tipo capa)
- Formatação numérica (% e 2 casas) nos indicadores e históricos
- Aba 'resumo_indicadores' com médias de ROE e Margem Líquida
- Gráfico automático 'Evolução Receita vs Lucro Líquido' na aba 'indicadores_hist'
- Autoajuste de largura das colunas nas planilhas geradas
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from datetime import date
from typing import List, Optional, Tuple, Dict, Any

import pandas as pd

try:
    from getDataCVM import DFP
except Exception as e:
    DFP = None  # type: ignore

# =========================
# Helpers de texto/dados
# =========================

def _norm(s: str) -> str:
    """Normaliza texto: remove acentos, caixa baixa, trim e colapsa espaços."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.lower().strip().split())


def _digits(s: str) -> str:
    return "".join(ch for ch in str(s) if ch.isdigit())


def _first_existing(candidates: Tuple[str, ...], df: pd.DataFrame) -> Optional[str]:
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _pick_date_col(df: pd.DataFrame) -> Optional[str]:
    """Escolhe a coluna de data mais provável nos DFPs."""
    prefs = (
        "DT_FIM_EXERC", "DT_FIM_EXERCICIO", "DT_FIM",
        "DT_REFER", "DT_REFERENCIA", "DATA",
        "DT_INI_EXERC"
    )
    return _first_existing(prefs, df)


def _info(msg: str):
    print(f"[INFO] {msg}")


def _warn(msg: str):
    print(f"[WARN] {msg}", file=sys.stderr)


def _print_table(title: str, rows: List[Tuple[Any, ...]], headers: List[str] | None = None):
    print("\\n" + title)
    print("-" * len(title))
    if headers:
        print(" | ".join(headers))
        print("-" * (sum(len(h) for h in headers) + 3 * (len(headers) - 1)))
    for r in rows:
        print(" | ".join(map(str, r)))
    print("")


# =========================
# Padrões de contas
# =========================

# DRE
LUCRO_PATTERNS = [
    r"lucro liquido",
    r"lucro.*periodo",
    r"resultado liquido.*exercicio",
]

RECEITA_PATTERNS = [
    r"receita operacional liquida",
    r"receita liquida",
    r"receitas de vendas",
]

LUCRO_BRUTO_PATTERNS = [
    r"lucro bruto",
]

EBIT_PATTERNS = [
    r"resultado operacional",
    r"resultado antes.*receitas.*despesas financeiras",  # fallback aproximado
    r"resultado antes.*juros.*impostos",                 # EBIT (quando existir)
]

# Bancos (DRE)
RECEITA_PATTERNS_BANK = [
    r"(resultado|receita) de intermediacao financeira",
    r"receitas da intermediacao financeira",
]
LUCRO_PATTERNS_BANK = LUCRO_PATTERNS  # em geral funciona igual

# BPP
PL_PATTERNS = [
    r"patrimonio liquido",
    r"patrim[oô]nio liquido",
]

PC_PATTERNS = [
    r"passivo circulante(?!.*nao)",
    r"passivo circulante$",
]

PNC_PATTERNS = [
    r"passivo nao circulante",
    r"passivo n[aã]o circulante",
]

ATIVO_TOTAL_PATTERNS = [
    r"ativo total",
]

ATIVO_CIRC_PATTERNS = [
    r"ativo circulante$",
]

ESTOQUES_PATTERNS = [
    r"estoques?$",
]


def _looks_like_bank(name: str) -> bool:
    if not name:
        return False
    n = _norm(name)
    return ("banco" in n) or ("instituicao financeira" in n)


# =========================
# Carregamento de dados
# =========================

def _load_dfp_range(dfp: Any, dataset: str, start: int, end: int) -> pd.DataFrame:
    """Carrega um dataset com fallback para *_ind se *_con vier vazio."""
    try:
        df = pd.DataFrame(dfp.get_data(dataset, start=start, end=end))
    except Exception as e:
        _warn(f"Falha ao carregar {dataset}: {e}")
        df = pd.DataFrame()
    if df.empty and dataset.endswith("_con"):
        alt = dataset.replace("_con", "_ind")
        _info(f"{dataset} vazio. Tentando fallback {alt}...")
        try:
            df = pd.DataFrame(dfp.get_data(alt, start=start, end=end))
        except Exception as e:
            _warn(f"Falha no fallback {alt}: {e}")
            df = pd.DataFrame()
    return df


def _annotate_common(df: pd.DataFrame) -> pd.DataFrame:
    """Adiciona colunas auxiliares: nome normalizado, CNPJ digits, coluna de data escolhida."""
    if df is None or df.empty:
        return df
    df = df.copy()
    denom = _first_existing(("DENOM_CIA", "NOME_EMPRESA", "NM_CIA"), df) or ""
    cnpj_col = _first_existing(("CNPJ_CIA", "CNPJ_CIA_X", "CNPJ_CIA_Y"), df) or ""
    df["__empresa_norm"] = df.get(denom, "").astype(str).map(_norm)
    df["__cnpj_digits"] = df.get(cnpj_col, "").astype(str).map(_digits)
    date_col = _pick_date_col(df)
    if date_col:
        df["__datecol"] = date_col
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    else:
        df["__datecol"] = ""
    return df


# =========================
# Busca de empresas
# =========================

def search_companies(dre: pd.DataFrame, bpp: pd.DataFrame, query: str, limit: int = 30, exact: bool = False) -> pd.DataFrame:
    pool = pd.concat([dre, bpp], ignore_index=True)
    if pool.empty:
        return pd.DataFrame()

    denom = _first_existing(("DENOM_CIA", "NOME_EMPRESA", "NM_CIA"), pool) or ""
    cnpj_col = _first_existing(("CNPJ_CIA",), pool)
    cd_cvm = _first_existing(("CD_CVM",), pool)

    qnorm = _norm(query)
    qdigits = _digits(query)

    if qdigits:
        mask = pool["__cnpj_digits"].str.contains(qdigits, na=False)
    else:
        if exact:
            mask = (pool["__empresa_norm"] == qnorm)
        else:
            tokens = [t for t in qnorm.split() if len(t) >= 3]
            if tokens:
                # exige presença de todos os tokens em qualquer ordem
                pat = "".join([f"(?=.*\\b{re.escape(t)}\\b)" for t in tokens])
                mask = pool["__empresa_norm"].str.contains(pat, na=False, regex=True)
            else:
                mask = pool["__empresa_norm"].str.contains(re.escape(qnorm), na=False, regex=True)

    found = pool.loc[mask, [c for c in (denom, cnpj_col, cd_cvm) if c]].drop_duplicates()

    # rank simples por tamanho da string e presença exata
    def _rank(row):
        name = str(row.get(denom, ""))
        n = _norm(name)
        score = 0
        if exact and n == qnorm:
            score += 100
        score += -abs(len(n) - len(qnorm))
        return score

    if not found.empty:
        found = found.assign(__rank=found.apply(_rank, axis=1)).sort_values("__rank", ascending=False)
        found = found.drop(columns="__rank")

    return found.head(limit).reset_index(drop=True)


# =========================
# Alinhamento DRE x BPP
# =========================

def align_latest_common(dre_emp: pd.DataFrame, bpp_emp: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, Optional[pd.Timestamp]]:
    if dre_emp is None or bpp_emp is None or dre_emp.empty or bpp_emp.empty:
        return dre_emp, bpp_emp, None

    dre_date_col = dre_emp["__datecol"].iloc[0]
    bpp_date_col = bpp_emp["__datecol"].iloc[0]

    dre_emp = dre_emp.copy()
    bpp_emp = bpp_emp.copy()

    if dre_date_col:
        dre_emp["__year"] = pd.to_datetime(dre_emp[dre_date_col], errors="coerce").dt.year
    else:
        dre_emp["__year"] = pd.NA

    if bpp_date_col:
        bpp_emp["__year"] = pd.to_datetime(bpp_emp[bpp_date_col], errors="coerce").dt.year
    else:
        bpp_emp["__year"] = pd.NA

    dre_years = sorted([int(y) for y in pd.Series(dre_emp["__year"]).dropna().unique()])
    bpp_years = sorted([int(y) for y in pd.Series(bpp_emp["__year"]).dropna().unique()])
    common = sorted(set(dre_years) & set(bpp_years))

    if common:
        y = common[-1]
        dre_sub = dre_emp[dre_emp["__year"] == y].copy()
        bpp_sub = bpp_emp[bpp_emp["__year"] == y].copy()
        ref_date = None
        if dre_date_col and bpp_date_col:
            try:
                ref_date = min(dre_sub[dre_date_col].max(), bpp_sub[bpp_date_col].max())
            except Exception:
                ref_date = None
        return dre_sub, bpp_sub, ref_date

    # Fallback: usa o ano máximo de cada um e pega o menor
    if dre_years and bpp_years:
        anchor_year = min(max(dre_years), max(bpp_years))
        dre_sub = dre_emp[dre_emp["__year"] <= anchor_year].copy()
        bpp_sub = bpp_emp[bpp_emp["__year"] <= anchor_year].copy()
        ref_date = None
        if dre_date_col and bpp_date_col:
            try:
                ref_date = min(dre_sub[dre_date_col].max(), bpp_sub[bpp_date_col].max())
            except Exception:
                ref_date = None
        return dre_sub, bpp_sub, ref_date

    return dre_emp, bpp_emp, None


# =========================
# Extração de contas
# =========================

def _find_first_value(df: pd.DataFrame, patterns: List[str],
                      value_cols: Tuple[str, ...] = ("VL_CONTA", "VL_CONTA_OR")) -> Tuple[Optional[float], Optional[str], Optional[pd.DataFrame]]:
    """Procura pela primeira conta que casa com um dos padrões. Retorna (valor, nome_conta, amostra_linhas)."""
    if df is None or df.empty:
        return None, None, None

    conta_col = _first_existing(("CONTA_DFP", "DS_CONTA", "CONTA", "DS_CONTA_COLUNA", "DS_CONTA_DEM", "DESC_CONTA"), df)
    if not conta_col:
        return None, None, None

    val_col = _first_existing(value_cols, df)
    if not val_col:
        return None, None, None

    tmp = df.copy()
    tmp["__conta_norm"] = tmp[conta_col].astype(str).map(_norm)
    for pat in patterns:
        m = tmp["__conta_norm"].str.contains(pat, na=False, regex=True)
        if m.any():
            sample = tmp.loc[m, [conta_col, val_col]].copy()
            vals = pd.to_numeric(tmp.loc[m, val_col], errors="coerce")
            if vals.notna().any():
                idx = vals[vals.notna()].index[-1]
                return float(vals.loc[idx]), str(tmp.loc[idx, conta_col]), sample
    return None, None, None


def _extract_series_by_year(df: pd.DataFrame, patterns: List[str]) -> pd.Series:
    """
    Retorna uma série indexada por ano com o último valor encontrado para cada ano que casar com os padrões.
    """
    if df is None or df.empty:
        return pd.Series(dtype="float64")

    conta_col = _first_existing(("CONTA_DFP", "DS_CONTA", "CONTA", "DS_CONTA_COLUNA", "DS_CONTA_DEM", "DESC_CONTA"), df)
    val_col = _first_existing(("VL_CONTA", "VL_CONTA_OR"), df)
    datecol = df["__datecol"].iloc[0] if "__datecol" in df.columns and not df.empty else None
    if not (conta_col and val_col and datecol):
        return pd.Series(dtype="float64")

    tmp = df.copy()
    tmp["__year"] = pd.to_datetime(tmp[datecol], errors="coerce").dt.year
    tmp["__conta_norm"] = tmp[conta_col].astype(str).map(_norm)

    mask = False
    for pat in patterns:
        m = tmp["__conta_norm"].str.contains(pat, na=False, regex=True)
        mask = m if isinstance(mask, bool) else (mask | m)

    sub = tmp.loc[mask, ["__year", val_col]].dropna()
    if sub.empty:
        return pd.Series(dtype="float64")

    # Para cada ano, pega o último valor reportado
    sub = sub.sort_values(["__year"]).dropna()
    out = sub.groupby("__year")[val_col].last()
    out = pd.to_numeric(out, errors="coerce")
    out = out.dropna()
    out.index = out.index.astype(int)
    return out


# =========================
# Excel helpers
# =========================

def _format_for_excel(hist: pd.DataFrame) -> pd.DataFrame:
    """Formata numéricos: percentuais em %, demais com 2 casas."""
    if hist is None or hist.empty:
        return hist
    hist = hist.copy()
    num_cols = hist.select_dtypes(include="number").columns
    for c in num_cols:
        if any(k in c for k in ["margem", "roe", "roa", "crescimento_"]):
            hist[c] = (hist[c] * 100.0).round(2)
        else:
            hist[c] = hist[c].round(2)
    return hist


def _autosize_columns(ws):
    """Ajusta a largura das colunas com base no conteúdo (openpyxl)."""
    try:
        from openpyxl.utils import get_column_letter
    except Exception:
        return
    for col_cells in ws.columns:
        length = 0
        col = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            length = max(length, len(val))
        ws.column_dimensions[col].width = min(max(12, length + 2), 60)


def _add_chart_receita_lucro(ws, hist_len: int):
    """Adiciona gráfico Receita vs Lucro Líquido na planilha indicadores_hist (openpyxl)."""
    try:
        from openpyxl.chart import LineChart, Reference
        chart = LineChart()
        chart.title = "Evolução Receita vs Lucro Líquido"
        chart.y_axis.title = "Valores (R$)"
        chart.x_axis.title = "Ano"

        # Assumindo layout: col A=ano, B=receita, C=lucro_liquido
        data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=hist_len+1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=hist_len+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "N2")
    except Exception as e:
        _warn(f"Não foi possível inserir o gráfico (openpyxl): {e}")


# =========================
# Execução principal
# =========================

def _fmt(v: Optional[float]) -> str:
    return "NA" if v is None else f"{v:,.2f}"


def _compute_historical_indicators(dre_emp: pd.DataFrame, bpp_emp: pd.DataFrame, is_bank: bool) -> pd.DataFrame:
    """Calcula indicadores por ano usando as séries extraídas."""
    # Séries DRE
    lucro_liq_s = _extract_series_by_year(dre_emp, LUCRO_PATTERNS_BANK if is_bank else LUCRO_PATTERNS)
    receita_s   = _extract_series_by_year(dre_emp, RECEITA_PATTERNS_BANK if is_bank else RECEITA_PATTERNS)
    lucro_bruto_s = _extract_series_by_year(dre_emp, LUCRO_BRUTO_PATTERNS)
    ebit_s        = _extract_series_by_year(dre_emp, EBIT_PATTERNS)

    # Séries BPP
    pl_s       = _extract_series_by_year(bpp_emp, PL_PATTERNS)
    pc_s       = _extract_series_by_year(bpp_emp, PC_PATTERNS)
    pnc_s      = _extract_series_by_year(bpp_emp, PNC_PATTERNS)
    ativo_tot_s= _extract_series_by_year(bpp_emp, ATIVO_TOTAL_PATTERNS)
    ac_s       = _extract_series_by_year(bpp_emp, ATIVO_CIRC_PATTERNS)
    estoques_s = _extract_series_by_year(bpp_emp, ESTOQUES_PATTERNS)

    years = sorted(set(lucro_liq_s.index) | set(receita_s.index) | set(pl_s.index) |
                   set(pc_s.index) | set(pnc_s.index) | set(ativo_tot_s.index) |
                   set(ac_s.index) | set(estoques_s.index) | set(lucro_bruto_s.index) |
                   set(ebit_s.index))

    rows = []
    for y in years:
        lucro = float(lucro_liq_s.get(y)) if y in lucro_liq_s.index else None
        rec   = float(receita_s.get(y)) if y in receita_s.index else None
        pl    = float(pl_s.get(y)) if y in pl_s.index else None
        pc    = float(pc_s.get(y)) if y in pc_s.index else None
        pnc   = float(pnc_s.get(y)) if y in pnc_s.index else None
        ativo = float(ativo_tot_s.get(y)) if y in ativo_tot_s.index else None
        ac    = float(ac_s.get(y)) if y in ac_s.index else None
        est   = float(estoques_s.get(y)) if y in estoques_s.index else None
        bruto = float(lucro_bruto_s.get(y)) if y in lucro_bruto_s.index else None
        ebit  = float(ebit_s.get(y)) if y in ebit_s.index else None

        divida = None
        if (pc is not None) or (pnc is not None):
            divida = (pc or 0.0) + (pnc or 0.0)

        # Indicadores
        roe = (lucro / pl) if (lucro is not None and pl not in (None, 0)) else None
        roa = (lucro / ativo) if (lucro is not None and ativo not in (None, 0)) else None
        margem_liq = (lucro / rec) if (lucro is not None and rec not in (None, 0)) else None
        margem_bruta = (bruto / rec) if (bruto is not None and rec not in (None, 0)) else None
        margem_oper = (ebit / rec) if (ebit is not None and rec not in (None, 0)) else None
        endivid = (divida / ativo) if (divida is not None and ativo not in (None, 0)) else None
        comp_div = (pc / divida) if (pc is not None and divida not in (None, 0)) else None
        liq_corr = (ac / pc) if (ac not in (None, 0) and pc not in (None, 0)) else None
        liq_seca = ((None if ac is None else (ac - (est or 0.0))) / pc) if (pc not in (None, 0) and ac is not None) else None

        rows.append({
            "ano": y,
            "receita": rec,
            "lucro_liquido": lucro,
            "lucro_bruto": bruto,
            "ebit": ebit,
            "pl": pl,
            "pc": pc,
            "pnc": pnc,
            "divida_bruta": divida,
            "ativo_total": ativo,
            "ativo_circulante": ac,
            "estoques": est,
            # Indicadores
            "roe": roe,
            "roa": roa,
            "margem_liquida": margem_liq,
            "margem_bruta": margem_bruta,
            "margem_operacional": margem_oper,
            "endividamento_sobre_ativo": endivid,
            "composicao_divida_pc_sobre_divida": comp_div,
            "liquidez_corrente": liq_corr,
            "liquidez_seca": liq_seca,
        })

    hist = pd.DataFrame(rows).sort_values("ano").reset_index(drop=True)

    # Crescimentos YoY
    for col in ["receita", "lucro_liquido"]:
        if col in hist.columns:
            hist[f"crescimento_{col}_yoy"] = hist[col].pct_change()

    return hist


def cmd_search(args):
    if DFP is None:
        _warn("Biblioteca getDataCVM não encontrada. Instale com: pip install getDataCVM")
        sys.exit(1)

    dfp = DFP()
    ini = args.ini or 2015
    fim = args.fim or date.today().year

    _info(f"Carregando DRE e BPP ({ini}-{fim}) para busca...")
    dre = _load_dfp_range(dfp, "dre_con", ini, fim)
    bpp = _load_dfp_range(dfp, "bpp_con", ini, fim)
    dre = _annotate_common(dre)
    bpp = _annotate_common(bpp)

    out = search_companies(dre, bpp, args.q, limit=args.limit, exact=args.exact)
    if out.empty:
        _warn("Nada encontrado. Tente variar o nome, usar CNPJ ou remover --exact.")
        return

    _print_table("Resultados da busca",
                 [tuple(x) for x in out.itertuples(index=False)],
                 list(out.columns))


def cmd_analyze(args):
    if DFP is None:
        _warn("Biblioteca getDataCVM não encontrada. Instale com: pip install getDataCVM")
        sys.exit(1)

    if not args.empresa and not args.cnpj:
        _warn("Informe --empresa ou --cnpj.")
        sys.exit(2)

    dfp = DFP()
    ini = args.ini or 2018
    fim = args.fim or date.today().year

    _info(f"Carregando DRE {ini}-{fim}...")
    dre = _annotate_common(_load_dfp_range(dfp, "dre_con", ini, fim))
    _info(f"Carregando BPP {ini}-{fim}...")
    bpp = _annotate_common(_load_dfp_range(dfp, "bpp_con", ini, fim))

    if dre.empty or bpp.empty:
        _warn("DRE ou BPP veio vazio. Verifique intervalo de anos ou conexão.")

    # Filtra empresa
    denom = _first_existing(("DENOM_CIA", "NOME_EMPRESA", "NM_CIA"), dre if not dre.empty else bpp) or ""
    cnpj_col = _first_existing(("CNPJ_CIA",), dre if not dre.empty else bpp)

    target_name = args.empresa or ""
    target_cnpj = _digits(args.cnpj or "") if args.cnpj else ""

    if target_cnpj:
        dre_emp = dre[dre["__cnpj_digits"] == target_cnpj].copy()
        bpp_emp = bpp[bpp["__cnpj_digits"] == target_cnpj].copy()
        alvo = target_cnpj
    else:
        qnorm = _norm(target_name)
        if args.exact:
            m_dre = dre["__empresa_norm"] == qnorm
            m_bpp = bpp["__empresa_norm"] == qnorm
        else:
            tokens = [t for t in qnorm.split() if len(t) >= 3]
            if tokens:
                pat = "".join([f"(?=.*\\b{re.escape(t)}\\b)" for t in tokens])
                m_dre = dre["__empresa_norm"].str.contains(pat, na=False, regex=True)
                m_bpp = bpp["__empresa_norm"].str.contains(pat, na=False, regex=True)
            else:
                m_dre = dre["__empresa_norm"].str.contains(re.escape(qnorm), na=False, regex=True)
                m_bpp = bpp["__empresa_norm"].str.contains(re.escape(qnorm), na=False, regex=True)

        dre_emp = dre[m_dre].copy()
        bpp_emp = bpp[m_bpp].copy()
        alvo = target_name

    if dre_emp.empty and bpp_emp.empty:
        _warn("Empresa não encontrada. Tente --exact, variações de nome ou use --cnpj.")
        sys.exit(3)

    # Alinha por ano (para relatório do último ano)
    dre_ult, bpp_ult, ref_date = align_latest_common(dre_emp, bpp_emp)

    # Detecta se é banco
    name_for_detect = ""
    if denom:
        if not dre_emp.empty:
            name_for_detect = str(dre_emp[denom].dropna().iloc[0])
        elif not bpp_emp.empty:
            name_for_detect = str(bpp_emp[denom].dropna().iloc[0])
    is_bank = _looks_like_bank(name_for_detect or alvo)

    # --- Métricas do último ano (compatibilidade com versão anterior)
    lucro_liq, lucro_conta, lucro_sample = _find_first_value(dre_ult, LUCRO_PATTERNS_BANK if is_bank else LUCRO_PATTERNS)
    receita_liq, receita_conta, receita_sample = _find_first_value(dre_ult, RECEITA_PATTERNS_BANK if is_bank else RECEITA_PATTERNS)
    pl, pl_conta, pl_sample = _find_first_value(bpp_ult, PL_PATTERNS)
    pc, pc_conta, pc_sample = _find_first_value(bpp_ult, PC_PATTERNS)
    pnc, pnc_conta, pnc_sample = _find_first_value(bpp_ult, PNC_PATTERNS)
    ativo_total, ativo_total_conta, _ = _find_first_value(bpp_ult, ATIVO_TOTAL_PATTERNS)
    ativo_circ, ativo_circ_conta, _ = _find_first_value(bpp_ult, ATIVO_CIRC_PATTERNS)
    estoques, estoques_conta, _ = _find_first_value(bpp_ult, ESTOQUES_PATTERNS)

    divida_bruta = ((pc or 0.0) + (pnc or 0.0)) if (pc is not None or pnc is not None) else None

    if args.debug_accounts:
        _info(f"Contas casadas: lucro='{lucro_conta}', receita='{receita_conta}', PL='{pl_conta}', PC='{pc_conta}', PNC='{pnc_conta}', ATIVO_TOTAL='{ativo_total_conta}', ATIVO_CIRC='{ativo_circ_conta}', ESTOQUES='{estoques_conta}'")

    # Tabela principal (último ano)
    rows = [
        ("Data Ref.", ref_date.date().isoformat() if ref_date else "-", "-"),
        ("Lucro Líquido", _fmt(lucro_liq), lucro_conta or "-"),
        ("Receita", _fmt(receita_liq), receita_conta or "-"),
        ("Lucro Bruto", _fmt(_find_first_value(dre_ult, LUCRO_BRUTO_PATTERNS)[0]), "lucro bruto (detecção automática)"),
        ("EBIT (aprox.)", _fmt(_find_first_value(dre_ult, EBIT_PATTERNS)[0]), "resultado operacional/EBIT (aprox.)"),
        ("Patrimônio Líquido", _fmt(pl), pl_conta or "-"),
        ("Ativo Total", _fmt(ativo_total), ativo_total_conta or "-"),
        ("Ativo Circulante", _fmt(ativo_circ), ativo_circ_conta or "-"),
        ("Estoques", _fmt(estoques), estoques_conta or "-"),
        ("Passivo Circulante", _fmt(pc), pc_conta or "-"),
        ("Passivo Não Circulante", _fmt(pnc), pnc_conta or "-"),
        ("Dívida Bruta", _fmt(divida_bruta), "-"),
    ]
    _print_table("Métricas (último ano comum entre DRE e BPP)", rows, ["Campo", "Valor", "Fonte (conta encontrada)"])

    # Warnings úteis
    for name, val in [("lucro", lucro_liq), ("receita", receita_liq), ("PL", pl), ("ativo total", ativo_total), ("ativo circulante", ativo_circ)]:
        if val is None:
            _warn(f"Conta de {name} não encontrada pelos padrões do setor. Use --samples para ver amostras e ajustar padrões.")

    # Indicadores (último ano)
    indicadores: Dict[str, Optional[float]] = {}
    if pl not in (None, 0) and lucro_liq is not None:
        indicadores["roe"] = float(lucro_liq / pl)
    if ativo_total not in (None, 0) and lucro_liq is not None:
        indicadores["roa"] = float(lucro_liq / ativo_total)
    if (receita_liq not in (None, 0)) and (lucro_liq is not None):
        indicadores["margem_liquida"] = float(lucro_liq / receita_liq)
    # bruto/operacional
    lucro_bruto_ult = _find_first_value(dre_ult, LUCRO_BRUTO_PATTERNS)[0]
    if (receita_liq not in (None, 0)) and (lucro_bruto_ult is not None):
        indicadores["margem_bruta"] = float(lucro_bruto_ult / receita_liq)
    ebit_ult = _find_first_value(dre_ult, EBIT_PATTERNS)[0]
    if (receita_liq not in (None, 0)) and (ebit_ult is not None):
        indicadores["margem_operacional"] = float(ebit_ult / receita_liq)
    if (ativo_total not in (None, 0)) and (divida_bruta is not None):
        indicadores["endividamento_sobre_ativo"] = float(divida_bruta / ativo_total)
    if (divida_bruta not in (None, 0)) and (pc is not None):
        indicadores["composicao_divida_pc_sobre_divida"] = float(pc / divida_bruta)
    if (pc not in (None, 0)) and (ativo_circ not in (None, 0)):
        indicadores["liquidez_corrente"] = float(ativo_circ / pc)
    if (pc not in (None, 0)) and (ativo_circ is not None):
        indicadores["liquidez_seca"] = float(((ativo_circ or 0.0) - (estoques or 0.0)) / pc)

    if indicadores:
        _print_table("Indicadores (último ano)", [(k, f"{v:.4f}") for k, v in indicadores.items()], ["Indicador", "Valor"])

    # Séries históricas completas
    hist = _compute_historical_indicators(dre_emp, bpp_emp, is_bank)
    if not hist.empty:
        _info("Séries históricas calculadas (indicadores_hist).")
        # Preview
        try:
            _print_table("Histórico (prévia)", [tuple(x) for x in hist.head(5).itertuples(index=False)], list(hist.columns))
        except Exception:
            pass

    # Samples (opcional)
    if args.samples:
        def _show_sample(lbl, sample):
            if isinstance(sample, pd.DataFrame) and not sample.empty:
                print(f"\\nAmostras - {lbl}")
                print(sample.head(10).to_string(index=False))
        _show_sample("Lucro Líquido", _find_first_value(dre_ult, LUCRO_PATTERNS_BANK if is_bank else LUCRO_PATTERNS)[2])
        _show_sample("Receita", _find_first_value(dre_ult, RECEITA_PATTERNS_BANK if is_bank else RECEITA_PATTERNS)[2])
        _show_sample("PL", _find_first_value(bpp_ult, PL_PATTERNS)[2])
        _show_sample("PC", _find_first_value(bpp_ult, PC_PATTERNS)[2])
        _show_sample("PNC", _find_first_value(bpp_ult, PNC_PATTERNS)[2])

    # Payload para export
    payload: Dict[str, Any] = {
        "alvo": alvo,
        "is_bank": is_bank,
        "periodo": {"ini": ini, "fim": fim},
        "referencia": ref_date.date().isoformat() if ref_date else None,
        "metricas": {
            "lucro_liquido": lucro_liq,
            "receita": receita_liq,
            "lucro_bruto": lucro_bruto_ult,
            "ebit": ebit_ult,
            "patrimonio_liquido": pl,
            "ativo_total": ativo_total,
            "ativo_circulante": ativo_circ,
            "estoques": estoques,
            "passivo_circulante": pc,
            "passivo_nao_circulante": pnc,
            "divida_bruta": divida_bruta,
        },
        "contas_encontradas": {
            "lucro_liquido": lucro_conta,
            "receita": receita_conta,
            "patrimonio_liquido": pl_conta,
            "ativo_total": ativo_total_conta,
            "ativo_circulante": ativo_circ_conta,
            "estoques": estoques_conta,
            "passivo_circulante": pc_conta,
            "passivo_nao_circulante": pnc_conta,
        },
        "indicadores": indicadores,
    }

    # ===== Exportação =====
    if args.out:
        out = args.out
        if out.lower().endswith(".json"):
            with open(out, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            _info(f"Exportado: {out}")
        elif out.lower().endswith((".xlsx", ".xls")):
            # xls e xlsx via ExcelWriter; preferível .xlsx
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                # 1) RESUMO limpo (capa)
                resumo_df = pd.DataFrame([
                    ["Empresa", payload["alvo"]],
                    ["Período", f"{payload['periodo']['ini']}–{payload['periodo']['fim']}"],
                    ["Referência", payload["referencia"]],
                    ["É banco?", "Sim" if payload["is_bank"] else "Não"],
                ], columns=["Campo", "Valor"])
                resumo_df.to_excel(writer, sheet_name="resumo", index=False)

                # 2) DADOS achatados (referência rápida)
                base = []
                for k, v in payload["metricas"].items():
                    base.append(("metrica", k, v))
                for k, v in payload["indicadores"].items():
                    base.append(("indicador", k, v))
                dados_df = pd.DataFrame(base, columns=["tipo", "chave", "valor"])
                dados_df.to_excel(writer, sheet_name="dados", index=False)

                # 3) Abas completas (todos os anos) para auditoria ampla
                def _drop_aux_cols(df):
                    aux = ["__empresa_norm", "__cnpj_digits", "__datecol", "__year"]
                    cols = [c for c in df.columns if c not in aux]
                    return df[cols].copy()

                if isinstance(dre_emp, pd.DataFrame) and not dre_emp.empty:
                    _drop_aux_cols(dre_emp).to_excel(writer, sheet_name="DRE_full", index=False)
                if isinstance(bpp_emp, pd.DataFrame) and not bpp_emp.empty:
                    _drop_aux_cols(bpp_emp).to_excel(writer, sheet_name="BPP_full", index=False)

                # 4) Recorte do último ano
                ref_year = None
                try:
                    if isinstance(dre_ult, pd.DataFrame) and not dre_ult.empty and "__year" in dre_ult.columns:
                        ref_year = int(dre_ult["__year"].dropna().iloc[0])
                except Exception:
                    ref_year = None

                if isinstance(dre_ult, pd.DataFrame) and not dre_ult.empty:
                    sheet = f"DRE_ref{('_' + str(ref_year)) if ref_year else ''}"
                    _drop_aux_cols(dre_ult).to_excel(writer, sheet_name=sheet, index=False)
                if isinstance(bpp_ult, pd.DataFrame) and not bpp_ult.empty:
                    sheet = f"BPP_ref{('_' + str(ref_year)) if ref_year else ''}"
                    _drop_aux_cols(bpp_ult).to_excel(writer, sheet_name=sheet, index=False)

                # 5) Indicadores históricos com formatação
                hist_fmt = _format_for_excel(hist) if isinstance(hist, pd.DataFrame) and not hist.empty else pd.DataFrame()
                if not hist_fmt.empty:
                    hist_fmt.to_excel(writer, sheet_name="indicadores_hist", index=False)

                    # 6) Contas históricas base
                    contas_hist = hist_fmt[[
                        "ano", "receita", "lucro_liquido", "lucro_bruto", "ebit", "pl", "pc", "pnc",
                        "divida_bruta", "ativo_total", "ativo_circulante", "estoques"
                    ]].copy()
                    contas_hist.to_excel(writer, sheet_name="contas_hist", index=False)

                    # 7) Resumo de indicadores (médias)
                    media = hist_fmt[["roe", "margem_liquida"]].mean(numeric_only=True)
                    media_df = pd.DataFrame({
                        "Indicador": ["ROE médio (%)", "Margem Líquida média (%)"],
                        "Valor": media.round(2).values
                    })
                    media_df.to_excel(writer, sheet_name="resumo_indicadores", index=False)

                # === Pós-processamento openpyxl: autoajuste + gráfico ===
                try:
                    wb = writer.book
                    # Autoajuste
                    for name in ["resumo", "dados", "DRE_full", "BPP_full",
                                 f"DRE_ref{('_' + str(ref_year)) if ref_year else ''}",
                                 f"BPP_ref{('_' + str(ref_year)) if ref_year else ''}",
                                 "indicadores_hist", "contas_hist", "resumo_indicadores"]:
                        if name in wb.sheetnames:
                            _autosize_columns(wb[name])

                    # Gráfico na indicadores_hist
                    if "indicadores_hist" in wb.sheetnames and not hist_fmt.empty:
                        ws = wb["indicadores_hist"]
                        _add_chart_receita_lucro(ws, hist_len=len(hist_fmt))

                except Exception as e:
                    _warn(f"Pós-processamento do Excel falhou: {e}")

            _info(f"Exportado: {out}")
        else:
            _warn("Extensão não reconhecida. Use .json ou .xlsx/.xls")

    return payload


def build_parser():
    p = argparse.ArgumentParser(description="Análise via getDataCVM (DFP)")
    sub = p.add_subparsers(dest="cmd", required=True)

    # search
    p_s = sub.add_parser("search", help="Buscar empresas por nome ou CNPJ")
    p_s.add_argument("--q", required=True, help="Nome (ou parte) ou CNPJ")
    p_s.add_argument("--ini", type=int, default=2015)
    p_s.add_argument("--fim", type=int, default=date.today().year)
    p_s.add_argument("--limit", type=int, default=30)
    p_s.add_argument("--exact", action="store_true", help="Casamento exato do nome normalizado")
    p_s.set_defaults(func=cmd_search)

    # analyze
    p_a = sub.add_parser("analyze", help="Analisar uma empresa")
    p_a.add_argument("--empresa", help="Nome da empresa (DENOM_CIA)")
    p_a.add_argument("--cnpj", help="CNPJ da companhia")
    p_a.add_argument("--ini", type=int, default=2018)
    p_a.add_argument("--fim", type=int, default=date.today().year)
    p_a.add_argument("--out", help="Arquivo de saída (.json ou .xlsx)")
    p_a.add_argument("--exact", action="store_true", help="Casamento exato do nome (quando usar --empresa)")
    p_a.add_argument("--debug-accounts", action="store_true", help="Exibe as contas que casaram para cada métrica")
    p_a.add_argument("--samples", action="store_true", help="Mostra amostras das linhas que casaram (debug)")
    p_a.set_defaults(func=cmd_analyze)

    return p


def main(argv=None):
    parser = build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    main()
